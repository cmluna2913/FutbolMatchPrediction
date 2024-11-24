%load_ext autoreload

%autoreload 2

import configparser
import os
import time
import random
import pandas as pd
import requests
import openpyxl
import sqlite3

from io import StringIO
from pathlib import Path
from bs4 import BeautifulSoup
from alive_progress import alive_bar

# some stuff I set up in a config file so I don't have to keep updating certain
# variables in every script
config = configparser.ConfigParser()
config.read('../src/config.ini')

config = configparser.ConfigParser()
config.read('../src/config.ini')

# the output path is specified in the config.ini file
output = Path(config['paths']['output'])
# I want data for the 2022 through 2024 season
yearly_directories = [Path(output/f"mls_{year}") for year in range(2022,2025)]

# create output directory and sub-directories if doesnt exist
for directory in yearly_directories+[output]:
    try:
        assert directory.exists()
    except:
        os.mkdir(directory)
        
# I will be web-scraping alot, so I made this function as a result

def get_html_data(url, parser='html.parser') -> BeautifulSoup:
    '''
    Extract html data from specified url and return a bs4 object.
    Parser can be specified if needed. Default is html.parser.
    '''
    response = requests.get(url)
    html_content = response.text
    soup = BeautifulSoup(html_content, parser)
    
    return soup

# I ended up using this a lot in the end
def get_table_data_from_html(soup) -> list:
    '''
    Extract tables from bs4 object and return a list of dataframes.
    '''
    
    # get all tables in the html
    tables = soup.findAll('table')

    # create dfs for each table and append each one to a list
    dfs_from_tables = []
    for table in tables:
        dfs_from_tables.append(pd.read_html(StringIO(str(table)))[0])
    
    return dfs_from_tables

def get_all_player_match_data(excel_file_path) -> tuple[pd.DataFrame, list]:
    '''
    Input an excel file of the copied tables from the websites and return a tuple
    where the first element is a dataframe of all player data and the second element
    is a list of URLs where extraction failed. This attempts to extract data from
    the failed URLs once more before completing.
    '''
    # get year from directory the excel file is in
    year = os.path.dirname(excel_file_path).split('\\')[-1].replace('mls_', '')
    # extract all links from the excel file
    # returned as a df in case of future use with an idea I had
    player_links = get_all_player_match_data_links(excel_file_path, year)
    # initialize empty df for all player data
    player_data_df = pd.DataFrame()
    
    # generate the initial player data df and failed links list
    player_data_df, failed_links = generate_player_df(list(player_links['stat_link']), player_data_df)
    
    # if first round has failed links, retry them
    # sometimes extracting the data the first time just doesn't work but rerunning usually does
    if len(failed_links)>0:
        player_data_df, failed_links = generate_player_df(failed_links, player_data_df)

    correct_col_names = ['Date', 'Day', 'Comp', 'Round', 'Venue', 'Result', 'Squad', 'Opponent',
       'Start', 'Pos', 'Min', 'Gls', 'Ast', 'PK', 'PKatt', 'Sh', 'SoT', 'CrdY',
       'CrdR', 'Touches', 'Tkl', 'Int', 'Blocks', 'xG', 'npxG', 'xAG', 'SCA',
       'GCA', 'Cmp', 'Att', 'Cmp%', 'PrgP', 'Carries', 'PrgC', 'Att_TakeOn', 'Succ',
       'Match Report', 'player_url']
    player_data_df.columns = correct_col_names

    return player_data_df, failed_links 

def generate_player_df(player_links, df=pd.DataFrame()) -> tuple[pd.DataFrame, list]:
    '''
    Given an iterable of individual player links (and a df to modify), return a tuple (DataFrame, list)
    where the dataframe contains all player data and the list contains URLs with failed extractions.
    '''
    # get total number of players for alive_bar
    total_players = len(player_links)
    # to append failed links to a list
    failed_links = []
    
    # progress bar, force_tty=True might be needed depending if animations don't show for you
    # I used jupyter notebooks, so I needed this
    with alive_bar(total_players, force_tty=True) as bar:
        for player_url in player_links:
            # limited to 10 requests a minute per website rules :(
            # I set this as 7 seconds just to be safe but you can adjust as needed
            # in the config file
            time.sleep(int(config['other']['request_time_limit']))
            
            # if you get a player df, append it to df, otherwise add to failed links
            temp_df = attempt_data_extraction(player_url)
            if type(temp_df)==type(None):
                failed_links.append(player_url)
            else:
                temp_df['player'] = player_url
                df = pd.concat([df, temp_df], ignore_index=True)
            bar()
    # return df and failed links
    return df, failed_links

def attempt_data_extraction(url) -> pd.DataFrame | None:
    '''
    Attempt to get a player data dataframe. Return the df if successful,
    otherwise return None.
    '''
    # try to get a player_df and return it
    try:
        player_df = get_player_data_df(url)
        return player_df
    # otherwise, let the user know and return nothing
    except:
        print(f"Could not get player data for {url}")
        return None

def get_all_player_match_data_links(excel_file_path, year) -> pd.DataFrame:
    '''
    Extract all links from the excel that is a copy of the website data. Return
    a dataframe of the player names and URLs.
    '''
    # read excel file
    all_players = pd.read_excel(excel_file_path)
    wb = openpyxl.load_workbook(excel_file_path)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    # get hyper links from file
    # in this situation, links didn't start until +2 and are found in column 37
    # this may need to be adjusted depending on how the data comes out
    all_players['stat_link'] = [ws.cell(row=i+2, column=37).hyperlink.target for i in range(all_players.shape[0])]
    # save all links out
    all_players[['Player', 'stat_link']].to_csv(output/f'mls_{year}' / 'player_links.csv', index=False)
    
    # return df of just the player name and stat link
    return all_players[['Player', 'stat_link']]
    
    
def get_player_data_df(url) -> pd.DataFrame:
    '''
    Using the URL for the individual player, get the first table on the website
    and return it as a dataframe.
    '''
    # get html data from url
    player_data_html = get_html_data(url)
    # generate tables list
    player_df = get_table_data_from_html(player_data_html)
    # add player url to data in first df in list
    # the lists always had 1 df but would break unless I left it as-is
    # player_df[0]['player'] = player_url
    
    return player_df[0]

# this url gives me a list of all players in the current league
base_url = 'https://fbref.com/en/comps/22/Major-League-Soccer-Stats'

# this page gives me a bunch of tables for team stats in the current moment
# not sure how much of the data here will be useful, but I'll grab it just-in-case
# html = get_html_data(base_url + 'players/')

# team_stat_dfs = get_table_data_from_html(html)

player_data_2024_df, failed_links_2024 = get_all_player_match_data(config['paths']['all_players_2024'])
player_data_2024_df
player_data_2024_df.to_csv(output/"mls_2024"/"all_player_data_2024.csv")

player_data_2023_df, failed_links_2023 = get_all_player_match_data(config['paths']['all_players_2023'])
player_data_2023_df
player_data_2023_df.to_csv(output/"mls_2023"/"all_player_data_2023.csv")

player_data_2022_df, failed_links_2022 = get_all_player_match_data(config['paths']['all_players_2022'])
player_data_2022_df
player_data_2022_df.to_csv(output/"mls_2022"/"all_player_data_2022.csv")